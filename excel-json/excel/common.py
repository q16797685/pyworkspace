#!/bin/python3
# -*- encoding:utf-8 -*-


import json


import openpyxl
from collections import defaultdict

data = openpyxl.load_workbook("e:\\demo\\department.xlsx")
sheet = data['1']


def merge():
    # TODO 查询该sheet表单所有合并单元格
    merge_lists = sheet.merged_cells
    # print(merge_lists)
    merge_all_list = []
    # TODO 遍历合并单元格
    for merge_list in merge_lists:
        # TODO 获取单个合并单元格的起始行(row)和起始列(col)
        row_min, row_max, col_min, col_max = merge_list.min_row, merge_list.max_row, merge_list.min_col, merge_list.max_col
        if (row_min != row_max and col_min != col_max):
            row_col = [(x, y) for x in range(row_min, row_max+1) for y in range(col_min, col_max+1)]
            merge_all_list.append(row_col)
        elif (row_min==row_max and col_min != col_max):
            row_col = [(row_min, y) for y in range(col_min, col_max + 1)]
            merge_all_list.append(row_col)
        elif (row_min != row_max and col_min == col_max):
            row_col = [(x, col_min) for x in range(row_min, row_max + 1)]
            merge_all_list.append(row_col)
    return merge_all_list
    # TODO 得到的是个这样的列表值：[[(2, 1), (3, 1)], [(10, 1), (10, 2), (10, 3), (11, 1), (11, 2), (11, 3)]]


def merge_values(*rr):
    mm_list = merge()
    for ii in range(0, len(mm_list)):
        if rr in mm_list[ii]:
            value11 = sheet.cell(row=mm_list[ii][0][0], column=mm_list[ii][0][1]).value
            return value11
    else:
        value2 = sheet.cell(*rr).value
        # if value2 is None:
        #     return "None：单元格无数据"
        # else:
        return value2


def list_excel_data():
    list_val = []
    for x in range(sheet.min_row, sheet.max_row+1):
        if x != 1:
            dict_val = {}
            for y in range(sheet.min_column, sheet.max_column+1):
                kk = (1, y)
                i = (x, y)
                dict_val[merge_values(*kk)] = merge_values(*i)
            list_val.append(dict_val)
    return list_val



all_data = list_excel_data()
excel_json = json.dumps(all_data,ensure_ascii=False)
print(all_data) # 列表
for i in all_data:  # 遍历列表，得出字典类型
    for akey,avalue in i.items(): # 遍历字典，得出key值和values值
        print(akey,avalue)
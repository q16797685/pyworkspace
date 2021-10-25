#!/bin/python3
# -*- encoding:utf-8 -*-


import openpyxl
from interfaceTest.base.base_path import *
from interfaceTest.base.base_path import _testcase_path


class HandleExcel:
    def __init__(self, file_name=None, sheet_name=None):

        if file_name:
            self.file_path = os.path.join(_testcase_path, file_name)
            self.sheet_name = sheet_name
        else:
            self.file_path = os.path.join(_testcase_path, 'workstation.xlsx')
            self.sheet_name = 'workstation'

        self.wb = openpyxl.load_workbook(self.file_path)
        self.sheet = self.wb[self.sheet_name]
        self.ncoles = self.sheet.max_column
        self.nrows = self.sheet.max_row

    def cell_value(self, row=1, column=1):
        return self.sheet.cell(row, column).value

    def get_title(self):
        title = []
        for column in range(1, self.ncoles+1):
            title.append(self.cell_value(1, column))
        return title

    def get_excel_data(self):
        finally_data = []
        for row in range(2,self.nrows):
            result_dict = {}
            for column in range(1,self.ncoles+1):
                result_dict[self.get_title()[column-1]] = self.cell_value(row, column)
            finally_data.append(result_dict)
        print(finally_data[0]['parameter'])
        return finally_data



if __name__ == '__main__':
    r = HandleExcel()
    print(r.get_excel_data())
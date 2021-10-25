#!/bin/python3
# -*- encoding:utf-8 -*-



import openpyxl


import os
from openpyxl import load_workbook

# data = openpyxl.load_workbook(r"e:\\demo\\de.xlsx")
# sheet = data['生产管理部']

wk = openpyxl.load_workbook(r"e:\\demo\\d1.xlsx")  # 加载已经存在的excel
wk_name = wk.sheetnames
wk_sheet = wk[wk_name[0]]

wk_sheet.cell(row=1, column=2, value='业务域')  # 在第二行，第二列下入“大区”数值
wk_sheet.cell(row=1, column=3, value='业务组件')
wk_sheet.cell(row=1, column=4, value='系统')
wk_sheet.cell(row=1, column=5, value='功能点')
wk_sheet.cell(row=1, column=6, value='痛点')

wk.save(r"e:\\demo\\d1.xlsx")
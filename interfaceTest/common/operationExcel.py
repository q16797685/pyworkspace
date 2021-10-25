#!/bin/python3
# -*- encoding:utf-8 -*-


"""获取excel表数据"""


import openpyxl
import xlrd
import os
from interfaceTest.base.base_path import _testcase_path


class ExcelVarles:
    case_Id = '用例ID'
    case_Module = '用例模块'
    case_Name = '用例名称'
    case_Url = '接口地址'
    case_Method = '请求方法'
    case_Type = '请求类型'
    case_Headers = '请求头'
    case_Data = '请求参数'
    case_Code = '状态码'
    case_Result = '期望结果'


class OperationExcel:

    def __init__(self, file_name=None, sheet_name=None):

        if file_name:
            self.file_path = os.path.join(_testcase_path, file_name)
            self.sheet_name = sheet_name
        else:
            self.file_path = os.path.join(_testcase_path, 'workstation.xls')
            self.sheet_name = 'workstation'

        self.wb = xlrd.open_workbook(self.file_path)
        self.sheet = self.wb.sheet_by_index(0)

    def get_excel_data(self):
        data = []
        title = self.sheet.row_values(0)
        for row in range(1,self.sheet.nrows):
            row_value = self.sheet.row_values(row)
            data.append(dict(zip(title,row_value)))
        return data

    #     self.wb = openpyxl.load_workbook(self.file_path)
    #     self.sheet = self.wb[self.sheet_name]
    #     self.ncoles = self.sheet.max_column
    #     self.nrows = self.sheet.max_row
    #
    # def cell_value(self, row=1, column=1):
    #     return self.sheet.cell(row, column).value
    #
    # def get_title(self):
    #     title = []
    #     for column in range(1, self.ncoles + 1):
    #         title.append(self.cell_value(1, column))
    #     return title
    #
    # def get_excel_data(self):
    #     finally_data = []
    #     for row in range(2, self.nrows):
    #         result_dict = {}
    #         for column in range(1, self.ncoles + 1):
    #             result_dict[self.get_title()[column - 1]] = self.cell_value(row, column)
    #         finally_data.append(result_dict)
    #     print(finally_data[0]['parameter'])
    #     return finally_data


if __name__ == '__main__':
    r = OperationExcel()
    d = ExcelVarles()
    print(d.case_Headers)
    for i in r.get_excel_data():
        print(i['请求头'])
